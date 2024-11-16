import os
import logging
from openpyxl import load_workbook
from copy import copy
from datetime import datetime

# Configure logging
logging.basicConfig(level=logging.INFO)

def format_date(date_str):
    """Convert date string to DD/MM/YYYY format"""
    try:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
        return date_obj.strftime('%d/%m/%Y')
    except ValueError as e:
        logging.error(f"Date format error: {e}")
        return date_str

def generate_excel(school_centre, lecturer_name, designation, ic_number, course_details):
    try:
        # Load template
        template_path = os.path.join(os.path.abspath(os.path.dirname(__file__)), 
                                   "files", 
                                   "Part-Time Lecturer Requisition Form - template new.xlsx")
        output_folder = os.path.join(os.path.abspath(os.path.dirname(__file__)), "temp")
        output_filename = f"{lecturer_name}.xlsx"
        output_path = os.path.join(output_folder, output_filename)

        # Ensure output directory exists
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        # Load workbook
        template_wb = load_workbook(template_path)
        template_ws = template_wb.active

        # Insert lecturer details
        template_ws['C5'].value = school_centre
        template_ws['C6'].value = lecturer_name
        template_ws['C7'].value = designation
        template_ws['H6'].value = ic_number

        # Store first record template (A9:L22)
        first_record_template = []
        for row in range(9, 23):
            row_data = []
            for col in range(1, 13):  # A to L
                cell = template_ws.cell(row=row, column=col)
                row_data.append({
                    'value': cell.value,
                    'style': copy(cell._style),
                    # Get formula from cell.value if it starts with '='
                    'formula': cell.value if isinstance(cell.value, str) and cell.value.startswith('=') else None
                })
            first_record_template.append(row_data)

        # Track total cost cells for final sum
        total_cost_cells = ['I20']  # First record's total cost cell

        # Process each course
        for index, course in enumerate(course_details):
            if index == 0:
                # First record uses existing template structure
                insert_record(template_ws, course, 9)
            else:
                # Calculate insertion point
                insert_point = 23 + (14 * (index - 1))
                
                # Insert new rows
                template_ws.insert_rows(insert_point, 14)
                
                # Copy template structure and formulas
                copy_record_structure(template_ws, first_record_template, insert_point)
                
                # Insert course data
                insert_record(template_ws, course, insert_point)
                
                # Update formulas for this record
                update_record_formulas(template_ws, insert_point)
                
                # Track total cost cell
                total_cost_cells.append(f'I{insert_point + 11}')

        # Update final total cost formula
        final_total_row = 23 + (14 * (len(course_details) - 1))
        template_ws[f'I{final_total_row}'].value = f'=SUM({",".join(total_cost_cells)})'

        # Save the file
        template_wb.save(output_path)
        logging.info(f"Excel file generated successfully at: {output_path}")
        return output_filename

    except Exception as e:
        logging.error(f"Error generating Excel file: {e}")
        raise

def copy_record_structure(ws, template_data, start_row):
    """
    Copy the record structure from template to new rows
    
    Parameters:
    - ws: worksheet to modify
    - template_data: stored template structure (from first record)
    - start_row: starting row number for the new record
    """
    try:
        # Iterate through each row in the template
        for row_idx, row_data in enumerate(template_data):
            # Calculate target row
            target_row = start_row + row_idx
            
            # Copy each cell in the row
            for col_idx, cell_data in enumerate(row_data, start=1):
                # Get target cell
                target_cell = ws.cell(row=target_row, column=col_idx)
                
                # Copy static values (like labels and descriptions)
                if cell_data['value'] and not cell_data['formula']:
                    target_cell.value = cell_data['value']
                
                # Copy cell style (formatting)
                target_cell._style = copy(cell_data['style'])
                
        logging.info(f"Successfully copied record structure to row {start_row}")
        
    except Exception as e:
        logging.error(f"Error copying record structure: {e}")
        raise

def insert_record(ws, course, start_row):
    """
    Insert course details into the Excel worksheet
    
    Parameters:
    - ws: worksheet to modify
    - course: dictionary containing course details
    - start_row: starting row number for this record
    """
    try:
        # Calculate row positions relative to start_row
        subject_title_row = start_row + 1  # Row 10 for first record
        subject_level_row = start_row + 2  # Row 11 for first record
        teaching_period_row = start_row + 4  # Row 13 for first record
        
        # Category rows (Lecture, Tutorial, Practical, Blended)
        category_start = start_row + 6  # Row 15 for first record
        
        # Insert basic course information
        ws[f'C{subject_title_row}'].value = course['subject_title']
        ws[f'I{subject_title_row}'].value = course['subject_code']
        ws[f'C{subject_level_row}'].value = course['program_level']
        
        # Format and insert teaching period dates
        start_date = format_date(course['start_date'])
        end_date = format_date(course['end_date'])
        ws[f'C{teaching_period_row}'].value = f"From {start_date} to {end_date}"
        
        # Insert hours data
        ws[f'D{category_start}'].value = course['lecture_hours']
        ws[f'D{category_start + 1}'].value = course['tutorial_hours']
        ws[f'D{category_start + 2}'].value = course['blended_hours']
        ws[f'D{category_start + 3}'].value = course['practical_hours']
        
        # Insert weeks data
        ws[f'G{category_start}'].value = course['lecture_weeks']
        ws[f'G{category_start + 1}'].value = course['tutorial_weeks']
        ws[f'G{category_start + 2}'].value = course['elearning_weeks']
        ws[f'G{category_start + 3}'].value = course['practical_weeks']
        
        # Insert hourly rate
        ws[f'D{start_row + 11}'].value = course['hourly_rate']  # Row 20 for first record
        
        logging.info(f"Successfully inserted course data starting at row {start_row}")
        
    except Exception as e:
        logging.error(f"Error inserting record: {e}")
        raise

def update_record_formulas(ws, start_row):
    """
    Update Excel formulas for a record block
    
    Parameters:
    - ws: worksheet to modify
    - start_row: starting row number for this record
    """
    try:
        # Calculate formula positions
        category_start = start_row + 6  # Row 15 for first record (categories start)
        total_row = start_row + 11      # Row 20 for first record (totals row)
        
        # Update category total formulas (hours × weeks)
        # Example: For first record, updates I15 to I18
        for i in range(4):  # 4 categories: Lecture, Tutorial, Practical, Blended
            row = category_start + i
            # Formula: hours × weeks (e.g., =D15*G15)
            ws[f'I{row}'].value = f'=D{row}*G{row}'
        
        # Update total hours formula
        # Example: For first record, G20 = SUM(I15:I18)
        ws[f'G{total_row}'].value = f'=SUM(I{category_start}:I{category_start+3})'
        
        # Update total cost formula
        # Example: For first record, J20 = D20*G20
        ws[f'I{total_row}'].value = f'=D{total_row}*G{total_row}'
        
        logging.info(f"Successfully updated formulas starting at row {start_row}")
        
    except Exception as e:
        logging.error(f"Error updating record formulas: {e}")
        raise
